# License: GNU General Public License v3. See license.txt

from __future__ import unicode_literals
from datetime import datetime
import frappe
import os
import frappe.utils
from frappe.desk import query_report
from six import string_types
from openpyxl import load_workbook
import base64
import logging

logging.basicConfig(filename='scorecard.log', level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))

@frappe.whitelist()
def generate_scorecard(sales_person, outstanding_total, target):
    fiscal_year = frappe.defaults.get_user_default("fiscal_year") 
    fiscal_year_start =datetime.strptime(frappe.defaults.get_user_default("year_start_date"),"%Y-%m-%d")
    fiscal_year_end = datetime.strptime(frappe.defaults.get_user_default("year_end_date"),"%Y-%m-%d")

    # Gets current location of file and opens workbook
    wb = load_workbook(os.path.join(__location__, 'sales_scorecard.xlsx'))
    summary_sheet = wb['Summary']
    summary_sheet['A3'] = sales_person
    summary_sheet['C4'] = target
    if datetime.today().month in range(4,7):
        summary_sheet['B7'] = "=SUM('Invoicing This Quarter'!C:E)"
    elif datetime.today().month in range(7,10):
        summary_sheet['B7'] = "=SUM('Invoicing This Quarter'!F:H)"
    elif datetime.today().month in range(10,13):
        summary_sheet['B7'] = "=SUM('Invoicing This Quarter'!I:K)"
    elif datetime.today().month in range(1,4):
        summary_sheet['B7'] = "=SUM('Invoicing This Quarter'!C:E)"
    # Itterate through pending orders and write to Pending Sheet
    pending_orders = get_pending_orders(sales_person)
    pending_orders_sheet = wb['Pending Orders']
    for row in range(0,len(pending_orders)):
        i = 1
        for key in pending_orders[row].keys():
            _ = pending_orders_sheet.cell(column=i, row=row+2, value= pending_orders[row][key])
            i = i + 1
    
    # # get ageing as per sales person
    
    collections_sheet = wb['Pending Collections']
    filtered =filter(lambda result: result['sales_person'] == sales_person, outstanding_total)
    outstanding = list(filtered)
    for row in range(0,len(outstanding)):
        i = 1
        for key in outstanding[row].keys():
            _ = collections_sheet.cell(column=i, row=row+2, value=outstanding[row][key])
            i = i + 1
    
    # Get Monthly Invoicing
    invoicing = get_sales_data(sales_person, fiscal_year_start.year, fiscal_year_end.year)
    invoicing_sheet = wb['Invoicing This Quarter']
    for row in range(0,len(invoicing)):
        i = 1
        for key in invoicing[row].keys():
            _ = invoicing_sheet.cell(column=i, row=row+2, value= invoicing[row][key])
            i = i + 1
    
    wb.save(os.path.join(__location__, 'sales_scorecard_o.xlsx'))
    wb.close()
    
def get_sales_persons():
    #get all relevent Sales Persons
    sales_persons = frappe.db.sql(
				"""
				SELECT name, email_id, target
				FROM `tabSales Person`
				WHERE is_group = 0 AND enabled = 1
			""",
				as_dict=1,
			)
    
    return sales_persons

def get_pending_orders(sales_person):
    pending_orders = frappe.db.sql(
            """
            select 
            `tabSales Order`.`name` as "so_number",
            `tabSales Order`.`po_no` as "po_no",
            `tabSales Order`.`po_date` as "po_date",
            `tabSales Order`.`customer` as "customer",
            `tabSales Order`.`transaction_date` as "date",
            `tabSales Order Item`.item_code as "item_code",
            `tabSales Order Item`.description as "desctiption",
            `tabSales Order Item`.qty as "qty",
            `tabSales Order Item`.delivered_qty as "delivered_qty", 
            (`tabSales Order Item`.qty - ifnull(`tabSales Order Item`.delivered_qty, 0)) as "pending_qty",
            `tabSales Order Item`.base_rate as "rate",
            `tabSales Order Item`.amount as "amount",
            `tabSales Order Item`.gst_hsn_code as "hsn",
            (select SUM(`tabBin`.actual_qty) from tabBin WHERE tabBin.warehouse in (select name from tabWarehouse WHERE tabWarehouse.parent_warehouse = "WAGHOLI BIN RACK - DBTPL") and tabBin.item_code = `tabSales Order Item`.item_code) as "wagholi",
            (select SUM(`tabBin`.actual_qty) from tabBin WHERE tabBin.warehouse = "Pune - DBTPL" and tabBin.item_code = `tabSales Order Item`.item_code) as "pune",
            (select SUM(`tabBin`.actual_qty) from tabBin WHERE tabBin.warehouse in (select name from tabWarehouse WHERE tabWarehouse.parent_warehouse = "PIMPRI BIN RACK - DBTPL") and tabBin.item_code = `tabSales Order Item`.item_code) as "pimpri",
            (select SUM(`tabBin`.actual_qty) from tabBin WHERE tabBin.warehouse in (select name from tabWarehouse WHERE tabWarehouse.parent_warehouse = "MAHAPE BIN RACK - DBTPL") and tabBin.item_code = `tabSales Order Item`.item_code) as "mahape",
            `tabSales Order Item`.item_group as "item_group",
            `tabCustomer`.sales_person as "sales_person"
            from `tabSales Order` 
            JOIN `tabSales Order Item` on `tabSales Order Item`.`parent` = `tabSales Order`.`name` 
            JOIN `tabCustomer` on `tabCustomer`.`name` = `tabSales Order`.`customer`
            WHERE  `tabSales Order`.docstatus = 1  and `tabSales Order`.status not in ("Stopped", "Closed") and ifnull(`tabSales Order Item`.delivered_qty,0) < ifnull(`tabSales Order Item`.qty,0) and `tabCustomer`.sales_person = %s
            order by `tabSales Order`.transaction_date asc
        """,
            sales_person,
            as_dict=1,
        )
    
    return pending_orders

def get_sales_data(sales_person, start_year, end_year):
    invoicing = frappe.db.sql(
        """
        select 
        `tabCustomer`.name as "customer",
        `tabCustomer`.sales_person as "sales_person",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 4 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "april",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 5 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "may",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 6 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "june",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 7 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "july",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 8 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "august",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 9 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "september",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 10 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "october",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 11 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "november",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 12 and YEAR(`tabSales Invoice`.posting_date) = {start_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "december",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 1 and YEAR(`tabSales Invoice`.posting_date) = {end_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "january",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 2 and YEAR(`tabSales Invoice`.posting_date) = {end_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "february",
        (select SUM(`tabSales Invoice`.net_total) from `tabSales Invoice` WHERE MONTH(`tabSales Invoice`.posting_date) = 3 and YEAR(`tabSales Invoice`.posting_date) = {end_year} and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.customer = `tabCustomer`.name )as "march"
        from `tabCustomer` 
        WHERE `tabCustomer`.disabled = 0 and `tabCustomer`.sales_person = '{sales_person}' 
        """.format(sales_person = sales_person, start_year = start_year, end_year = end_year),
        as_dict = 1
    )
    return invoicing

def get_outstanding_data():
    outstanding = query_report.run(report_name = "Accounts Recc Summary", filters= {"company":"Dhupar Brothers Trading Pvt. Ltd.","report_date":datetime.today().strftime('%Y-%m-%d'),"ageing_based_on":"Posting Date","range1":30,"range2":60,"range3":90,"range4":120})
    outstanding = outstanding['result']
    outstanding.pop()
    sorted_outstanding = []
    cheques_on_hold = frappe.db.sql(
        """
        select 
        party , sum(paid_amount) as "amount"
        from `tabPayment Entry` 
        WHERE `tabPayment Entry`.workflow_state = "PDC" 
        and `tabPayment Entry`.payment_type = "Receive" 
        and `tabPayment Entry`.party_type = "Customer"
        and `tabPayment Entry`.posting_date < "{date}"
        GROUP BY party
        """.format(date = datetime.today().strftime('%Y-%m-%d')),
        as_dict = 1
    )
    cheques_in_hand = frappe.db.sql(
        """
        select 
        party , sum(paid_amount) as "amount"
        from `tabPayment Entry` 
        WHERE `tabPayment Entry`.workflow_state = "PDC" 
        and `tabPayment Entry`.payment_type = "Receive" 
        and `tabPayment Entry`.party_type = "Customer"
        and `tabPayment Entry`.posting_date > "{date}"
        GROUP BY party
        """.format(date = datetime.today().strftime('%Y-%m-%d')),
        as_dict = 1
    )

    for i in outstanding:

        sorted_outstanding.append(
            {
                'customer': i.party,
                'debtor_days': i.debtor_days,
                'cheques_on_hold': next((item["amount"] for item in cheques_on_hold if item["party"] == i.party), None),
                'cheques_in_hand': next((item["amount"] for item in cheques_in_hand if item["party"] == i.party), None),
                'advance': i.advance,
                'outstanding': i.outstanding,
                '0-30': i.range1,
                '31-60': i.range2,
                '61-90': i.range3,
                '91-120': i.range4,
                '121+': i.range5,
                'sales_person': i.sales_person if i.sales_person else "None"
            }
        )
    return sorted_outstanding

@frappe.whitelist()
def deliver_outstandings():
    total_outstanding = get_outstanding_data()

    sales_force = get_sales_persons()

    for i in sales_force:
        if i['email_id'] != None:
            try:
                generate_scorecard(i['name'], outstanding_total= total_outstanding, target= i['target'])

                message = """
                Please find the attached scorecard.<br>
                <br>
                Make sure the Pending Orders are LIVE.<br>
                Make sure your DSO is under 90 days.<br>
                <br>
                Best of Luck.
                """
                attachments = [{
                        'fname': "sales_scorecard_{date}.xlsx".format(date = datetime.today().strftime('%d-%m-%Y')),
                        'fcontent': open(os.path.join(__location__, 'sales_scorecard_o.xlsx'),'rb').read()
                    }]

                frappe.sendmail(
                        recipients = i['email_id'],
                        subject = "Your Daily Performance Scorecard",
                        message = message,
                        attachments = attachments
                    )
                frappe.db.commit()
            except Exception as e:
                pass

#  This function send the scorecard to haresh Patel of slaes person datta,bhushan,sahid
@frappe.whitelist()
def deliver_outstandings_to_haresh_of_sahid_data_bhushan():
    total_outstanding = get_outstanding_data()

    sales_force = frappe.db.sql(
                    """
                        SELECT name, email_id, target
                        FROM `tabSales Person`
                        WHERE
                        email_id in ('shahid.kureshi@dhuparbrothers.com','dattatray.mane@dhuparbrothers.com','bhushan@dhupargroup.com')
                        AND
                        is_group = 0 
                        AND 
                        enabled = 1
                    """,
                        as_dict=1,
                    )
    try:
        for i in sales_force:
            if i['email_id']:
                generate_scorecard(i['name'], outstanding_total= total_outstanding, target= i['target'])

                message = f"""
                <b>Sales Person Name : {i['name']}</b><br><br>
                Please find the attached scorecard.<br>
                <br>
                Make sure the Pending Orders are LIVE.<br>
                Make sure your DSO is under 90 days.<br>
                <br>
                Best of Luck.
                """
                attachments = [{
                        'fname': "sales_scorecard_{date}.xlsx".format(date = datetime.today().strftime('%d-%m-%Y')),
                        'fcontent': open(os.path.join(__location__, 'sales_scorecard_o.xlsx'),'rb').read()
                    }]

                frappe.sendmail(
                        recipients = 'haresh@dhuparbrothers.com',
                        subject = "Your Daily Performance Scorecard",
                        message = message,
                        attachments = attachments
                    )
                frappe.db.commit()
    except Exception as e:
        pass


# This function created when postal not working and shivaji sir want the scorecard
# shivaji Holkar

@frappe.whitelist()
def deliver_outstandings_to_shivaji():
    total_outstanding = get_outstanding_data()

    sales_force = frappe.db.sql(
                    """
                        SELECT name, email_id, target
                        FROM `tabSales Person`
                        WHERE
                        email_id in ('shivaji.holkar@dhuparbrothers.com')
                        AND
                        is_group = 0 
                        AND 
                        enabled = 1
                    """,
                        as_dict=1,
                    )

    for i in sales_force:
        if i['email_id']:
            generate_scorecard(i['name'], outstanding_total= total_outstanding, target= i['target'])

            message = f"""
            <b>Sales Person Name : {i['name']}</b><br><br>
            Please find the attached scorecard.<br>
            <br>
            Make sure the Pending Orders are LIVE.<br>
            Make sure your DSO is under 90 days.<br>
            <br>
            Best of Luck.
            """
            attachments = [{
                    'fname': "sales_scorecard_{date}.xlsx".format(date = datetime.today().strftime('%d-%m-%Y')),
                    'fcontent': open(os.path.join(__location__, 'sales_scorecard_o.xlsx'),'rb').read()
                }]

            frappe.sendmail(
                    recipients = 'shivaji.holkar@dhuparbrothers.com',
                    subject = "Your Daily Performance Scorecard",
                    message = message,
                    attachments = attachments
                )
            frappe.db.commit()
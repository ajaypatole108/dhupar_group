import frappe
import logging
from frappe.desk import query_report
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment
from openpyxl.styles.borders import Border,Side
from datetime import datetime
from frappe.utils.pdf import get_pdf
from frappe import _
# import pandas as pd

# logging.basicConfig(filename='send_ledger.log', level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))

@frappe.whitelist()
def send_ledger(**kwargs):
	# logging.info(kwargs)

	customer_name = kwargs.get('customer')
	from_date1 = datetime.strptime(kwargs.get('from_date'), "%Y-%m-%d")
	from_date = from_date1.strftime("%d-%m-%Y")

	to_date1 = datetime.strptime(kwargs.get('to_date'), "%Y-%m-%d")
	to_date = to_date1.strftime("%d-%m-%Y")

	email_id = kwargs.get('email_id')
	mobile = kwargs.get('mobile')

	report = query_report.run(report_name = "General Ledger",filters= {"company":"Dhupar Brothers Trading Pvt. Ltd.", "from_date": from_date1, "to_date": to_date1, "party_type":'Customer', "party": (customer_name,), "group_by": 'Group by Voucher (Consolidated)',"include_dimensions":True, "include_default_book_entries":True},ignore_prepared_report= True)
	ledger_report = report['result']
	# logging.info(ledger_report)
	# print(__location__)

	thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

	wb = load_workbook(os.path.join(__location__, 'ledger_statement_template.xlsx'))
	template_sheet = wb['Ledger']
	template_sheet['A5'] = customer_name

	template_sheet['C7'] = from_date
	template_sheet['E7'] = to_date

	opening = []
	credit = []
	debit = []
	gl_entry = []

	for i in ledger_report:
		if i['account'].replace("'","") == 'Opening':
			if i['balance'] > 0:
				template_sheet['D12'] = frappe.format_value(i['balance'], {"fieldtype":"Currency"})
				template_sheet['D12'].font = Font(bold=True,size=11)
				opening.append({'debit':abs(i['balance'])})
			else:
				template_sheet['E12'] = frappe.format_value(abs(i['balance']), {"fieldtype":"Currency"})
				template_sheet['E12'].font = Font(bold=True,size=11)
				opening.append({'credit':abs(i['balance'])})

		for k,v in i.items():
			if k == 'gl_entry':
				gl_entry.append(i['gl_entry'])
				if i['credit']:
					credit.append(i['credit'])
				if i['debit']:
					debit.append(i['debit'])
		cnt = 0
		for row in range(13,len(gl_entry)+13):
			# border code
			for i1 in range(row,row+1):
				for j1 in range(1,6):
					ch = chr(64+j1)
					template_sheet[ch+(str(i1))].border = thin_border

			for k,v in i.items():
				if k == 'gl_entry':
					if v == gl_entry[cnt]:
						_ = template_sheet.cell(column=1, row=row, value= i['posting_date'])
						template_sheet['A'+(str(row))].alignment = Alignment(horizontal='center')
						_ = template_sheet.cell(column=2, row=row, value= i['voucher_type'])
						template_sheet['B'+(str(row))].alignment = Alignment(horizontal='center')
						_ = template_sheet.cell(column=3, row=row, value= i['voucher_no'])
						template_sheet['C'+(str(row))].alignment = Alignment(horizontal='center')
						_ = template_sheet.cell(column=4, row=row, value= frappe.format_value(abs(i['debit']), {"fieldtype":"Currency"}))
						template_sheet['D'+(str(row))].alignment = Alignment(horizontal='center')
						_ = template_sheet.cell(column=5, row=row, value= frappe.format_value(abs(i['credit']), {"fieldtype":"Currency"}))
						template_sheet['E'+(str(row))].alignment = Alignment(horizontal='center')
				if i['account'].replace("'","") == 'Total':
					template_sheet['C'+(str(len(gl_entry)+13))].border = thin_border

					# border for last 3 rows
					for i1 in range(len(gl_entry)+13,len(gl_entry)+13+3):
						for j1 in range(1,6):
							ch = chr(64+j1)
							template_sheet[ch+(str(i1))].border = thin_border

					_ = template_sheet.cell(column=3, row=len(gl_entry)+13+1, value= 'Total')
					template_sheet['C'+(str(len(gl_entry)+13+1))].font = Font(bold=True,size=11)

					# cell1 = f"A{str(len(gl_entry)+13+1)}:C{str(len(gl_entry)+13+1)}"
					# template_sheet.merge_cells(f'{cell1}')
					# template_sheet.row_dimensions[row+1].height = 5

					for k,v in opening[0].items():
						if k == 'credit':
							_ = template_sheet.cell(column=4, row=len(gl_entry)+13+1, value= frappe.format_value(sum(debit),{"fieldtype":"Currency"}))
							template_sheet['D'+(str(len(gl_entry)+13+1))].font = Font(bold=True,size=11)
							_ = template_sheet.cell(column=5, row=len(gl_entry)+13+1, value= frappe.format_value(sum(credit)+v,{"fieldtype":"Currency"}))
							template_sheet['E'+(str(len(gl_entry)+13+1))].font = Font(bold=True,size=11)

						if k == 'debit':
							_ = template_sheet.cell(column=5, row=len(gl_entry)+13+1, value= frappe.format_value(sum(credit),{"fieldtype":"Currency"}))
							template_sheet['D'+(str(len(gl_entry)+13+1))].font = Font(bold=True,size=11)
							_ = template_sheet.cell(column=4, row=len(gl_entry)+13+1, value= frappe.format_value(sum(debit)+v,{"fieldtype":"Currency"}))
							template_sheet['E'+(str(len(gl_entry)+13+1))].font = Font(bold=True,size=11)

				if i['account'].replace("'","") == 'Closing (Opening + Total)':
					_ = template_sheet.cell(column=3, row=len(gl_entry)+13+2, value= 'Outstanding')
					template_sheet['C'+(str(len(gl_entry)+13+2))].font = Font(bold=True,size=11)
					_ = template_sheet.cell(column=5, row=len(gl_entry)+13+2, value= frappe.format_value(i['balance'],{"fieldtype":"Currency"}))
					template_sheet['E'+(str(len(gl_entry)+13+2))].font = Font(bold=True,size=11)
			cnt+=1

	# print(opening,gl_entry,credit,debit,len(gl_entry))

	wb.save(os.path.join(__location__, 'ledger_statement.xlsx'))
	wb.close()

	attachments = [{
					'fname': "ledger_statement.xlsx",
					'fcontent': open(os.path.join(__location__, 'ledger_statement.xlsx'),'rb').read()
				}]

	message = f"""
				Dear Sir/Madam,<br>
				<br>
				I trust this email finds you well. 
				we have attached your ledger statement for the period <b>{from_date} To {to_date}</b>.
				Please review the statement at your convenience and let us know if you have any questions or require clarification on any of the transactions.
				Our support team is available to assist you and can be reached at <b>accounts@dhuparbrothers.com</b>
				<br>
				<br>
				<br>
				Regards,
				<br>
				<br>
				Dhupar Brothers Trading Pvt. Ltd.
			"""

	frappe.sendmail(
						recipients = email_id,
						subject = "Ledger Statement",
						sender = "info@dhuparbrothers.com",
						message = message,
						attachments = attachments
					)
	frappe.db.commit()